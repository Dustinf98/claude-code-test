import os
import pandas as pd
from categorizer import categorize


def parse_cibc_csv(filepath):
    df = pd.read_csv(
        filepath,
        header=None,
        names=["Date", "Description", "Debit", "Credit", "Card"],
    )
    df = df[["Date", "Description", "Debit", "Credit"]].copy()
    df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m-%d")
    return df


def apply_categories(df, rules):
    df = df.copy()
    df["Category"] = df["Description"].apply(lambda d: categorize(d, rules))
    return df


def save_month(df, data_dir="data"):
    os.makedirs(data_dir, exist_ok=True)
    # Determine year-month from the Date column
    dates = pd.to_datetime(df["Date"])
    year_month = dates.dt.to_period("M").mode()[0]
    filename = os.path.join(data_dir, f"{year_month}.csv")
    df.to_csv(filename, index=False)
    return filename


def load_month(year_month, data_dir="data"):
    filename = os.path.join(data_dir, f"{year_month}.csv")
    if not os.path.exists(filename):
        return pd.DataFrame(columns=["Date", "Description", "Debit", "Credit", "Category"])
    return pd.read_csv(filename)


def load_all_months(data_dir="data"):
    if not os.path.exists(data_dir):
        return pd.DataFrame(columns=["Date", "Description", "Debit", "Credit", "Category"])
    frames = []
    for fname in sorted(os.listdir(data_dir)):
        if fname.endswith(".csv"):
            frames.append(pd.read_csv(os.path.join(data_dir, fname)))
    if not frames:
        return pd.DataFrame(columns=["Date", "Description", "Debit", "Credit", "Category"])
    return pd.concat(frames, ignore_index=True)


def list_months(data_dir="data"):
    if not os.path.exists(data_dir):
        return []
    months = []
    for fname in sorted(os.listdir(data_dir)):
        if fname.endswith(".csv") and not fname.endswith("_summary.csv"):
            months.append(fname.replace(".csv", ""))
    return months


def import_excel(filepath, data_dir="data"):
    """
    Reads Monthly Expenses.xlsx.
    If a sheet has Category + Total columns → treat as a pre-computed summary.
    Otherwise try to read as raw transactions.
    Returns list of (year_month, df) tuples that were saved.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    saved = []
    os.makedirs(data_dir, exist_ok=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        # Summary sheet detection: has "Category" and "Total" (or "Amount") columns
        lower_headers = [h.lower() for h in headers]
        if "category" in lower_headers and ("total" in lower_headers or "amount" in lower_headers):
            cat_idx = lower_headers.index("category")
            total_idx = lower_headers.index("total") if "total" in lower_headers else lower_headers.index("amount")
            data = []
            for row in rows[1:]:
                if row[cat_idx] is not None and row[total_idx] is not None:
                    data.append({"Category": row[cat_idx], "Total": float(row[total_idx])})
            if data:
                summary_df = pd.DataFrame(data)
                # Try to parse year-month from sheet name (e.g. "Oct 2024", "2024-10")
                ym = _parse_sheet_name_to_ym(sheet_name)
                out_path = os.path.join(data_dir, f"{ym}_summary.csv")
                summary_df.to_csv(out_path, index=False)
                saved.append((ym, summary_df))
        else:
            # Try raw transaction format
            try:
                df = pd.DataFrame(rows[1:], columns=headers)
                df.to_csv(os.path.join(data_dir, f"excel_{sheet_name}.csv"), index=False)
                saved.append((sheet_name, df))
            except Exception:
                pass

    return saved


def _parse_sheet_name_to_ym(name):
    import re
    from datetime import datetime
    # Try "YYYY-MM"
    m = re.match(r"(\d{4})-(\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Try "Mon YYYY" like "Oct 2024"
    for fmt in ("%b %Y", "%B %Y", "%b%Y", "%B%Y"):
        try:
            dt = datetime.strptime(name.strip(), fmt)
            return dt.strftime("%Y-%m")
        except ValueError:
            pass
    return name.replace(" ", "_")
